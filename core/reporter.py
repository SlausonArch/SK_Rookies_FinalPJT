"""
진단 결과 리포트 생성기 (Excel / Markdown / JSON / Log)
"""
import json
import os
from datetime import datetime
from core.result import ScanReport, Status, Severity

# 점검 ID 앞자리 → 섹션명 (Linux 기준)
_SECTION_NAMES: dict[str, str] = {
    "1": "계정 관리",
    "2": "파일 시스템",
    "3": "네트워크 서비스",
    "4": "로그 관리",
    "5": "주요 응용 설정",
    "6": "시스템 보안 설정",
    "7": "보안 패치",
}


REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")

# 상태/위험도별 색상 (Excel 셀 fill 색)
_STATUS_FILL = {
    "취약":   "FFEB9C9C",  # 연한 빨강
    "양호":   "FFC6EFCE",  # 연한 초록
    "수동점검": "FFFFEB9C",  # 연한 노랑
    "오류":   "FFD9D9D9",  # 회색
    "미해당": "FFEDEDED",  # 밝은 회색
}
_SEVERITY_FILL = {
    "위험": "FFFF0000",  # 빨강
    "높음": "FFFF6B6B",  # 연빨강
    "보통": "FFFFC000",  # 주황
    "낮음": "FFFFFF00",  # 노랑
    "정보": "FF9DC3E6",  # 하늘
}


def _ensure_dir():
    os.makedirs(REPORTS_DIR, exist_ok=True)


def _filename(report: ScanReport, ext: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_type = report.scan_type.replace("/", "_").replace(" ", "_")
    return os.path.join(REPORTS_DIR, f"{safe_type}_{ts}.{ext}")


# ── Excel ────────────────────────────────────────────────────────────

def save_excel(report: ScanReport) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    _ensure_dir()
    path = _filename(report, "xlsx")
    wb = Workbook()

    # ── 시트 1: 요약 ────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "요약"

    header_font  = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFFFF")
    header_fill  = PatternFill("solid", fgColor="FF2C3E50")
    title_font   = Font(name="맑은 고딕", bold=True, size=14)
    label_font   = Font(name="맑은 고딕", bold=True, size=10)
    value_font   = Font(name="맑은 고딕", size=10)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="FFBDC3C7")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    s = report.summary

    # 제목
    ws_sum["A1"] = "취약점 진단 결과 보고서"
    ws_sum["A1"].font = title_font
    ws_sum.merge_cells("A1:D1")
    ws_sum.row_dimensions[1].height = 30

    # 기본 정보
    meta = [
        ("진단 유형", report.scan_type),
        ("진단 대상", report.target),
        ("시작 시각", report.started_at),
        ("종료 시각", report.finished_at or "-"),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws_sum.cell(i, 1, label).font = label_font
        ws_sum.cell(i, 2, value).font = value_font
        ws_sum.row_dimensions[i].height = 18

    # 통계 표
    ws_sum["A8"] = "구분"
    ws_sum["B8"] = "건수"
    ws_sum["A8"].font = header_font
    ws_sum["B8"].font = header_font
    ws_sum["A8"].fill = header_fill
    ws_sum["B8"].fill = header_fill
    ws_sum["A8"].alignment = center
    ws_sum["B8"].alignment = center

    stat_rows = [
        ("전체",    s["total"],      "FFD9EAD3"),
        ("취약",    s["vulnerable"], "FFEB9C9C"),
        ("양호",    s["safe"],       "FFC6EFCE"),
        ("수동점검", s["manual"],    "FFFFEB9C"),
        ("오류",    s["error"],      "FFD9D9D9"),
        ("", "", ""),
        ("위험",    s["by_severity"]["위험"], "FFFF0000"),
        ("높음",    s["by_severity"]["높음"], "FFFF6B6B"),
        ("보통",    s["by_severity"]["보통"], "FFFFC000"),
        ("낮음",    s["by_severity"]["낮음"], "FFFFFF00"),
        ("정보",    s["by_severity"]["정보"], "FF9DC3E6"),
    ]
    for i, (label, val, color) in enumerate(stat_rows, start=9):
        if not label:
            continue
        c_label = ws_sum.cell(i, 1, label)
        c_val   = ws_sum.cell(i, 2, val)
        c_label.font = label_font
        c_label.alignment = center
        c_val.font   = Font(name="맑은 고딕", bold=True, size=12)
        c_val.alignment = center
        fill = PatternFill("solid", fgColor=color)
        c_label.fill = fill
        c_val.fill   = fill
        c_label.border = border
        c_val.border   = border
        ws_sum.row_dimensions[i].height = 22

    ws_sum.column_dimensions["A"].width = 14
    ws_sum.column_dimensions["B"].width = 10
    ws_sum.column_dimensions["C"].width = 30
    ws_sum.column_dimensions["D"].width = 30

    # ── 시트 2: 전체 결과 ────────────────────────────────────────────
    ws = wb.create_sheet("전체 결과")

    cols = ["ID", "카테고리", "점검 항목", "결과", "위험도",
            "점검 설명", "진단 결과 상세", "조치 방안", "근거", "점검 시각"]
    col_widths = [10, 16, 28, 10, 10, 36, 36, 36, 30, 18]

    for ci, (col, width) in enumerate(zip(cols, col_widths), start=1):
        cell = ws.cell(1, ci, col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, r in enumerate(report.results, start=2):
        row_data = [r.check_id, r.category, r.name, r.status.value,
                    r.severity.value, r.description, r.details,
                    r.recommendation, r.evidence, r.checked_at]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(ri, ci, val)
            cell.font      = value_font
            cell.alignment = left
            cell.border    = border

        # 결과 셀 색상
        status_cell   = ws.cell(ri, 4)
        severity_cell = ws.cell(ri, 5)
        status_fill   = _STATUS_FILL.get(r.status.value, "FFFFFFFF")
        severity_fill = _SEVERITY_FILL.get(r.severity.value, "FFFFFFFF")
        status_cell.fill   = PatternFill("solid", fgColor=status_fill)
        severity_cell.fill = PatternFill("solid", fgColor=severity_fill)
        status_cell.alignment   = center
        severity_cell.alignment = center

        ws.row_dimensions[ri].height = 40

    # ── 시트 3: 취약 항목만 ──────────────────────────────────────────
    ws_vuln = wb.create_sheet("취약 항목")
    for ci, (col, width) in enumerate(zip(cols, col_widths), start=1):
        cell = ws_vuln.cell(1, ci, col)
        cell.font      = header_font
        cell.fill      = PatternFill("solid", fgColor="FFC0392B")
        cell.alignment = center
        cell.border    = border
        ws_vuln.column_dimensions[get_column_letter(ci)].width = width
    ws_vuln.row_dimensions[1].height = 22
    ws_vuln.freeze_panes = "A2"

    vuln_rows = [r for r in report.results if r.status == Status.VULNERABLE]
    for ri, r in enumerate(vuln_rows, start=2):
        row_data = [r.check_id, r.category, r.name, r.status.value,
                    r.severity.value, r.description, r.details,
                    r.recommendation, r.evidence, r.checked_at]
        for ci, val in enumerate(row_data, start=1):
            cell = ws_vuln.cell(ri, ci, val)
            cell.font      = value_font
            cell.alignment = left
            cell.border    = border
        severity_cell = ws_vuln.cell(ri, 5)
        severity_cell.fill      = PatternFill("solid", fgColor=_SEVERITY_FILL.get(r.severity.value, "FFFFFFFF"))
        severity_cell.alignment = center
        ws_vuln.row_dimensions[ri].height = 40

    wb.save(path)
    return path


# ── Markdown ─────────────────────────────────────────────────────────

_STATUS_ICON = {
    "취약":    "🔴",
    "양호":    "🟢",
    "수동점검": "🟡",
    "오류":    "⚫",
    "미해당":  "⚪",
}
_SEVERITY_ICON = {
    "위험": "🚨",
    "높음": "🔴",
    "보통": "🟠",
    "낮음": "🟡",
    "정보": "🔵",
}


def save_markdown(report: ScanReport) -> str:
    _ensure_dir()
    path = _filename(report, "md")
    s = report.summary
    lines = []

    # 제목 & 메타
    lines += [
        f"# 취약점 진단 결과 — {report.scan_type}",
        "",
        f"| 항목 | 내용 |",
        f"|------|------|",
        f"| 진단 대상 | `{report.target}` |",
        f"| 시작 시각 | {report.started_at} |",
        f"| 종료 시각 | {report.finished_at or '-'} |",
        "",
    ]

    # 요약 통계
    lines += [
        "## 요약",
        "",
        "| 구분 | 건수 |",
        "|------|-----:|",
        f"| 전체 | **{s['total']}** |",
        f"| 🔴 취약 | **{s['vulnerable']}** |",
        f"| 🟢 양호 | {s['safe']} |",
        f"| 🟡 수동점검 | {s['manual']} |",
        f"| ⚫ 오류 | {s['error']} |",
        "",
        "| 위험도 | 건수 |",
        "|--------|-----:|",
        f"| 🚨 위험 | **{s['by_severity']['위험']}** |",
        f"| 🔴 높음 | **{s['by_severity']['높음']}** |",
        f"| 🟠 보통 | {s['by_severity']['보통']} |",
        f"| 🟡 낮음 | {s['by_severity']['낮음']} |",
        f"| 🔵 정보 | {s['by_severity']['정보']} |",
        "",
    ]

    # 취약 항목 상세 (우선 출력)
    vuln_items = [r for r in report.results if r.status == Status.VULNERABLE]
    if vuln_items:
        lines += ["## 취약 항목 상세", ""]
        for r in vuln_items:
            sev_icon = _SEVERITY_ICON.get(r.severity.value, "")
            lines += [
                f"### {sev_icon} [{r.check_id}] {r.name}",
                "",
                f"| | |",
                f"|---|---|",
                f"| 위험도 | {r.severity.value} |",
                f"| 카테고리 | {r.category} |",
                f"| 점검 설명 | {r.description} |",
                f"| 진단 결과 | {r.details} |",
                f"| 조치 방안 | {r.recommendation} |",
            ]
            if r.evidence:
                lines += [
                    "",
                    "**근거**",
                    "```",
                    r.evidence[:500],
                    "```",
                ]
            lines.append("")

    # 전체 결과 표
    lines += [
        "## 전체 점검 결과",
        "",
        "| ID | 점검 항목 | 결과 | 위험도 | 진단 결과 상세 |",
        "|----|-----------|:----:|:------:|----------------|",
    ]
    for r in report.results:
        st_icon  = _STATUS_ICON.get(r.status.value, "")
        sev_icon = _SEVERITY_ICON.get(r.severity.value, "")
        detail   = r.details.replace("\n", " ")[:60]
        lines.append(
            f"| `{r.check_id}` | {r.name} "
            f"| {st_icon} {r.status.value} "
            f"| {sev_icon} {r.severity.value} "
            f"| {detail} |"
        )

    lines.append("")
    content = "\n".join(lines)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    return path


# ── Log ──────────────────────────────────────────────────────────────

_STATUS_LABEL = {
    "취약":    "[취약]",
    "양호":    "[양호]",
    "수동점검": "[검토]",
    "오류":    "[검토]",
    "미해당":  "[ N/A]",
}

def save_log(report: ScanReport) -> str:
    """모든 점검 결과를 섹션별로 구분한 텍스트 로그 파일로 저장."""
    _ensure_dir()
    path = _filename(report, "log")
    s = report.summary
    W = 68

    lines: list[str] = []
    def hr(c="═"):
        lines.append(c * W)

    hr()
    lines.append(f"  취약점 진단 로그 | {report.scan_type}")
    hr()
    lines.append(f"  대상    : {report.target}")
    lines.append(f"  시작    : {report.started_at}")
    lines.append(f"  종료    : {report.finished_at or '-'}")
    hr("─")
    lines.append(f"  전체 {s['total']}건  |  "
                 f"[취약] {s['vulnerable']}  "
                 f"[양호] {s['safe']}  "
                 f"[검토] {s['manual'] + s['error']}  "
                 f"[N/A] {s['skipped']}")
    lines.append(f"  위험도 — 위험:{s['by_severity']['위험']}  "
                 f"높음:{s['by_severity']['높음']}  "
                 f"보통:{s['by_severity']['보통']}  "
                 f"낮음:{s['by_severity']['낮음']}")
    hr()
    lines.append("")

    # 섹션별로 그룹핑
    from collections import OrderedDict
    sections: dict[str, list] = OrderedDict()
    for r in report.results:
        sec = r.check_id.split(".")[0] if "." in r.check_id else "기타"
        sections.setdefault(sec, []).append(r)

    for sec_key, results in sections.items():
        sec_name = _SECTION_NAMES.get(sec_key, f"섹션 {sec_key}")
        lines.append(f"{'─' * W}")
        lines.append(f"  {sec_key}. {sec_name}")
        lines.append(f"{'─' * W}")
        for r in results:
            label = _STATUS_LABEL.get(r.status.value, "[    ]")
            lines.append(f"{label} [{r.check_id}] {r.name}")
            lines.append(f"  ▸ 위험도  : {r.severity.value}")
            lines.append(f"  ▸ 점검내용: {r.description}")
            lines.append(f"  ▸ 결  과  : {r.details}")
            lines.append(f"  ▸ 조치방안: {r.recommendation}")
            if r.evidence:
                evidence_lines = r.evidence.splitlines()
                lines.append(f"  ▸ 근  거  : {evidence_lines[0]}")
                for el in evidence_lines[1:5]:
                    lines.append(f"              {el}")
            lines.append("")

    hr()
    lines.append(f"  로그 생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    hr()

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


# ── JSON (내부 저장용) ────────────────────────────────────────────────

def save_json(report: ScanReport) -> str:
    _ensure_dir()
    path = _filename(report, "json")
    data = {
        "target": report.target,
        "scan_type": report.scan_type,
        "started_at": report.started_at,
        "finished_at": report.finished_at,
        "summary": report.summary,
        "results": [r.to_dict() for r in report.results],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path
